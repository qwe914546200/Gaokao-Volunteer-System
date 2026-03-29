import os
from functools import lru_cache
from fastapi import APIRouter, Query, HTTPException
from typing import Optional, List
from backend.database import get_db_connection

router = APIRouter()

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
RANKING_DIR = os.path.join(BASE_DIR, "数据资料", "全国高校排名数据")
MAJOR_SCORE_FILE = os.path.join(
    BASE_DIR,
    "数据资料",
    "全国高校录取位次和分数数据",
    "四川",
    "录取分数",
    "22-25年全国高校在四川的专业录取分数.xlsx"
)
BASIC_INFO_FILE = os.path.join(
    BASE_DIR,
    "数据资料",
    "全国高校目录和基本信息数据",
    "大学院校基础信息（含部标代码）.xlsx"
)

def normalize_subject(subject: str) -> str:
    if subject in ['物理类', '物理']:
        return '理科'
    if subject in ['历史类', '历史']:
        return '文科'
    return subject

def normalize_batch(batch: str) -> str:
    if not batch:
        return batch
    b = batch.strip().replace(' ', '')
    if '专科' in b:
        return '专科批'
    if '本科' in b:
        return '本科批'
    return b

def batch_like_value(batch: str) -> str:
    if batch == '本科批':
        return '本科%'
    if batch == '专科批':
        return '专科%'
    return batch

def is_dual_class_value(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return s != '' and s != '-'

@lru_cache(maxsize=1)
def load_school_website_map():
    website_map = {}
    try:
        from openpyxl import load_workbook
        if not os.path.exists(BASIC_INFO_FILE):
            return website_map
        wb = load_workbook(BASIC_INFO_FILE, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        idx_name = headers.index('中文名字') if '中文名字' in headers else None
        idx_site = headers.index('官网') if '官网' in headers else None
        if idx_name is None or idx_site is None:
            wb.close()
            return website_map
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[idx_name]).strip() if row[idx_name] is not None else ''
            site = str(row[idx_site]).strip() if row[idx_site] is not None else ''
            if name:
                website_map[name] = site
        wb.close()
    except Exception:
        return website_map
    return website_map

@lru_cache(maxsize=1)
def load_major_requirement_rows():
    rows = []
    try:
        from openpyxl import load_workbook
        if not os.path.exists(MAJOR_SCORE_FILE):
            return rows
        wb = load_workbook(MAJOR_SCORE_FILE, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({
                "year": r[idx.get("年份")] if idx.get("年份") is not None else None,
                "school_name": str(r[idx.get("院校名称")]).strip() if idx.get("院校名称") is not None and r[idx.get("院校名称")] is not None else "",
                "subject_type": str(r[idx.get("科类")]).strip() if idx.get("科类") is not None and r[idx.get("科类")] is not None else "",
                "batch": str(r[idx.get("批次")]).strip() if idx.get("批次") is not None and r[idx.get("批次")] is not None else "",
                "subject_requirement": str(r[idx.get("选科要求")]).strip() if idx.get("选科要求") is not None and r[idx.get("选科要求")] is not None else "",
                "professional_group": str(r[idx.get("所属专业组")]).strip() if idx.get("所属专业组") is not None and r[idx.get("所属专业组")] is not None else "",
                "major_name": str(r[idx.get("专业")]).strip() if idx.get("专业") is not None and r[idx.get("专业")] is not None else "",
                "min_score": r[idx.get("最低分数")] if idx.get("最低分数") is not None else None,
                "min_rank": r[idx.get("最低位次")] if idx.get("最低位次") is not None else None,
            })
        wb.close()
    except Exception:
        return rows
    return rows

@router.get("/directory")
def get_school_directory(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    level: Optional[str] = None,      # 本科 / 专科
    type_: Optional[str] = None,      # 综合 / 理工 / 师范 等
    province: Optional[str] = None,   # 所在省份
    keyword: Optional[str] = None     # 校名关键字
):
    """模块四：高校目录页面接口"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = "SELECT * FROM school_info WHERE 1=1"
        params = []
        
        if level:
            query += " AND level = ?"
            params.append(level)
        if type_ and type_ != "全部":
            query += " AND school_type = ?"
            params.append(type_)
        if province and province != "全部":
            query += " AND province = ?"
            params.append(province)
        if keyword:
            query += " AND school_name LIKE ?"
            params.append(f"%{keyword}%")
            
        # 查询总数
        count_query = f"SELECT COUNT(*) as total FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        
        # 分页并按排名排序
        query += " ORDER BY rank ASC, id ASC LIMIT ? OFFSET ?"
        params.extend([size, (page - 1) * size])
        
        cursor.execute(query, params)
        items = [dict(row) for row in cursor.fetchall()]
        
        return {
            "status": "success",
            "total": total,
            "page": page,
            "size": size,
            "items": items
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/filter")
def filter_schools(
    total_score: int,
    subject_type: str,
    batch: str,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    level: Optional[str] = None,      # 985 / 211 / 双一流 / 普通本科 / 专科
    province: Optional[str] = None,
    type_: Optional[str] = None,
    rec_type: Optional[str] = None    # reach / safe / guarantee
):
    """模块三：院校筛选列表接口 (需结合用户成绩进行推荐类别筛选)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        mapped_subject = normalize_subject(subject_type)
        mapped_batch = normalize_batch(batch)
        batch_like = batch_like_value(mapped_batch)

        # 获取最新的录取年份
        cursor.execute("SELECT MAX(year) as max_year FROM school_admission WHERE subject_type = ? AND batch LIKE ?", (mapped_subject, batch_like))
        year_row = cursor.fetchone()
        if not year_row or not year_row['max_year']:
            cursor.execute("SELECT MAX(year) as max_year FROM school_admission WHERE subject_type = ? AND batch LIKE ?", (subject_type, batch_like))
            year_row = cursor.fetchone()
            query_subject = subject_type
        else:
            query_subject = mapped_subject

        adm_year = year_row['max_year'] if year_row and year_row['max_year'] else 2024
        
        # 构建联合查询
        query = """
            SELECT a.school_name, a.province, a.batch, a.min_score, a.min_rank,
                   i.school_type, i.is_985, i.is_211, i.dual_class, i.level
            FROM school_admission a
            LEFT JOIN school_info i ON a.school_name = i.school_name
            WHERE a.year = ? AND a.subject_type = ? AND a.batch LIKE ? AND a.min_score IS NOT NULL
        """
        params = [adm_year, query_subject, batch_like]
        
        if province and province != "全部":
            query += " AND a.province = ?"
            params.append(province)
        if type_ and type_ != "全部":
            query += " AND i.school_type = ?"
            params.append(type_)
            
        if level and level != "全部":
            if level == "985":
                query += " AND i.is_985 = 1"
            elif level == "211":
                query += " AND i.is_211 = 1"
            elif level == "双一流":
                query += " AND i.dual_class IS NOT NULL AND i.dual_class != '' AND i.dual_class != '-'"
            elif level == "普通本科":
                query += " AND i.level = '本科' AND i.is_985 = 0 AND i.is_211 = 0 AND (i.dual_class IS NULL OR i.dual_class = '' OR i.dual_class = '-')"
            elif level == "专科":
                query += " AND i.level = '专科'"
                
        # 处理冲稳保条件筛选
        if rec_type and rec_type != "全部":
            # 冲一冲：5 <= diff <= 20  -> min_score - score >= 5 AND min_score - score <= 20
            # 稳一稳：-5 <= diff <= 4
            # 保一保：diff <= -10
            if rec_type == 'reach':
                query += " AND (a.min_score - ?) BETWEEN 5 AND 20"
                params.append(total_score)
            elif rec_type == 'safe':
                query += " AND (a.min_score - ?) BETWEEN -5 AND 4"
                params.append(total_score)
            elif rec_type == 'guarantee':
                query += " AND (a.min_score - ?) <= -10"
                params.append(total_score)

        # 查询总数
        count_query = f"SELECT COUNT(*) as total FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        
        # 分页排序 (这里可以按分数倒序)
        query += " ORDER BY a.min_score DESC LIMIT ? OFFSET ?"
        params.extend([size, (page - 1) * size])
        
        cursor.execute(query, params)
        items = []
        for row in cursor.fetchall():
            row_dict = dict(row)
            # 计算对应的冲稳保类型和概率
            diff = row_dict['min_score'] - total_score
            rt = 'unknown'
            prob = 0
            if 5 <= diff <= 20:
                rt = 'reach'
                prob = int(40 - (diff - 5) * (20 / 15))
            elif -5 <= diff <= 4:
                rt = 'safe'
                prob = int(50 + (5 - diff) * (25 / 10))
            elif diff <= -10:
                rt = 'guarantee'
                prob = int(80 + (-diff - 6) * (15 / 14)) if diff <= -6 else 80
                
            row_dict['rec_type'] = rt
            row_dict['probability'] = prob
            items.append(row_dict)
            
        return {
            "status": "success",
            "total": total,
            "page": page,
            "size": size,
            "items": items
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/provinces")
def get_available_provinces(subject_type: str, batch: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        mapped_subject = normalize_subject(subject_type)
        mapped_batch = normalize_batch(batch)
        batch_like = batch_like_value(mapped_batch)
        cursor.execute(
            "SELECT DISTINCT province FROM school_admission WHERE subject_type = ? AND batch LIKE ? AND province IS NOT NULL AND province != '' ORDER BY province",
            (mapped_subject, batch_like)
        )
        provinces = [row["province"] for row in cursor.fetchall()]
        return {"status": "success", "items": provinces}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/detail")
def get_school_detail(
    school_name: str,
    subject_type: str,
    batch: Optional[str] = None
):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        mapped_subject = normalize_subject(subject_type)
        mapped_batch = normalize_batch(batch) if batch else None
        batch_like = batch_like_value(mapped_batch) if mapped_batch else "%"

        cursor.execute(
            "SELECT * FROM school_info WHERE school_name = ? LIMIT 1",
            (school_name,)
        )
        info_row = cursor.fetchone()
        school_info = dict(info_row) if info_row else {}

        cursor.execute(
            """
            SELECT year, batch, min_score, min_rank, province
            FROM school_admission
            WHERE school_name = ? AND subject_type = ? AND batch LIKE ?
            ORDER BY year DESC
            """,
            (school_name, mapped_subject, batch_like)
        )
        admission_history = [dict(r) for r in cursor.fetchall()]

        cursor.execute(
            """
            SELECT year, major_name, min_score, min_rank
            FROM major_admission
            WHERE school_name = ? AND subject_type = ? AND batch LIKE ?
            ORDER BY year DESC, min_score DESC
            LIMIT 80
            """,
            (school_name, mapped_subject, batch_like)
        )
        major_rows = [dict(r) for r in cursor.fetchall()]

        requirement_rows = []
        for r in load_major_requirement_rows():
            if r["school_name"] == school_name and normalize_subject(r["subject_type"]) == mapped_subject:
                if not mapped_batch or normalize_batch(r["batch"]) == mapped_batch:
                    requirement_rows.append(r)

        professional_groups = sorted({x["professional_group"] for x in requirement_rows if x["professional_group"]})
        subject_requirements = sorted({x["subject_requirement"] for x in requirement_rows if x["subject_requirement"]})

        site_map = load_school_website_map()
        official_site = site_map.get(school_name, "")

        return {
            "status": "success",
            "school_name": school_name,
            "school_info": school_info,
            "official_site": official_site,
            "admission_history": admission_history,
            "major_history": major_rows,
            "professional_groups": professional_groups,
            "subject_requirements": subject_requirements
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@lru_cache(maxsize=16)
def _load_ranking_rows(category: str, source: str):
    from openpyxl import load_workbook
    source_map = {
        ("学科排名", "ESI学科排名"): os.path.join(RANKING_DIR, "学科排名", "2025年5月ESI学科排名.xlsx"),
        ("学科排名", "中国大学软科学科排名"): os.path.join(RANKING_DIR, "学科排名", "2025中国大学软科排名.xlsx"),
        ("学校综合排名", "ESI"): os.path.join(RANKING_DIR, "大学综合排名", "ESI排名.xlsx"),
        ("学校综合排名", "U.S.News世界大学"): os.path.join(RANKING_DIR, "大学综合排名", "U.S.News世界大学排名.xlsx"),
        ("学校综合排名", "泰晤士"): os.path.join(RANKING_DIR, "大学综合排名", "泰晤士排名.xlsx"),
        ("学校综合排名", "校友会"): os.path.join(RANKING_DIR, "大学综合排名", "校友会排行榜.xlsx"),
    }
    fp = source_map.get((category, source))
    if not fp or not os.path.exists(fp):
        return []

    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []

    if category == "学科排名" and source == "ESI学科排名":
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if not r or not r[0]:
                continue
            rows.append({"rank": i, "school_name": str(r[0]).strip(), "subject_name": str(r[3]).strip() if len(r) > 3 and r[3] else ""})
    elif category == "学科排名" and source == "中国大学软科学科排名":
        for r in ws.iter_rows(min_row=4, values_only=True):
            if not r:
                continue
            rank = r[0]
            name = r[10] if len(r) > 10 else None
            if rank and name:
                try:
                    rank = int(rank)
                except Exception:
                    pass
                rows.append({"rank": rank, "school_name": str(name).strip(), "subject_name": ""})
    elif category == "学校综合排名" and source == "ESI":
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[1]:
                continue
            rows.append({"rank": r[0], "school_name": str(r[1]).strip(), "subject_name": ""})
    elif category == "学校综合排名" and source == "U.S.News世界大学":
        for r in ws.iter_rows(min_row=3, values_only=True):
            if not r or not r[1]:
                continue
            rows.append({"rank": r[0], "school_name": str(r[1]).strip(), "subject_name": ""})
    elif category == "学校综合排名" and source == "泰晤士":
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[1]:
                continue
            rows.append({"rank": r[0], "school_name": str(r[1]).strip(), "subject_name": ""})
    elif category == "学校综合排名" and source == "校友会":
        for r in ws.iter_rows(min_row=6, values_only=True):
            if not r or not r[1]:
                continue
            rows.append({"rank": r[0], "school_name": str(r[1]).strip(), "subject_name": ""})

    wb.close()
    return rows

@router.get("/ranking/options")
def ranking_options():
    return {
        "status": "success",
        "items": {
            "学科排名": ["ESI学科排名", "中国大学软科学科排名"],
            "学校综合排名": ["ESI", "U.S.News世界大学", "泰晤士", "校友会"]
        }
    }

@router.get("/ranking/list")
def ranking_list(
    category: str = Query("学校综合排名"),
    source: str = Query("ESI"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None
):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        rows = _load_ranking_rows(category, source)
        if keyword:
            rows = [x for x in rows if keyword in x["school_name"]]
        total = len(rows)
        start = (page - 1) * size
        end = start + size
        page_rows = rows[start:end]

        for item in page_rows:
            cursor.execute("SELECT province, school_type, level, is_985, is_211, dual_class FROM school_info WHERE school_name = ? LIMIT 1", (item["school_name"],))
            info = cursor.fetchone()
            if info:
                item["province"] = info["province"]
                item["school_type"] = info["school_type"]
                item["level"] = info["level"]
                item["is_985"] = info["is_985"]
                item["is_211"] = info["is_211"]
                item["dual_class"] = info["dual_class"]
            else:
                item["province"] = ""
                item["school_type"] = ""
                item["level"] = ""
                item["is_985"] = 0
                item["is_211"] = 0
                item["dual_class"] = ""
            item["category"] = category
            item["source"] = source

        return {"status": "success", "total": total, "page": page, "size": size, "items": page_rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
